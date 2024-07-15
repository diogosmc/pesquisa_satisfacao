import tkinter as tk
from tkinter import ttk
import time
import os
import smtplib
import imaplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import load_workbook, Workbook
from datetime import datetime
from openpyxl.cell.cell import MergedCell

attachment_paths = []

def send_email():

    try:
        smtp_s = "smtp-mail.outlook.com"
        smtp_p = 587
        
        sender_email = "@hotmail.com"  # Substitua pelo seu email 
        sender_password = "senha"  # Substitua pela sua senha 
        recipient_email = email_entry.get()

        # Coletar respostas da GUI
        npedido = npedido_entry.get()
        Nome_Contato = entrevistado_entry.get()
        Ordem_Compra = OC_entry.get()
        telefone_user  = telefone_entry.get()
        cliente_user = cliente_entry.get()
        user_email = email_entry.get()
        attendance_rating = attendance_combobox.get()
        product_satisfaction = product_combobox.get()
        recommendation = recommendation_combobox.get()
        feedback = feedback_combobox.get()
        current_datetime = datetime.now().strftime("%d-%m-%Y %H:%M:%S")
        observaçao = obs_entry.get()

        # Preparar o corpo do email
        email_body = f"""\
        <html>
        <body>
        <b>Pesquisa de Satisfação</b><br>
        <b>Data/Hora</b>: {current_datetime}<br>
        <b>Email</b>: {user_email}<br>
        ___________________________________________________________________________________________________________________
        <p><b>De 0 à 10, qual nota você atribui ao ATENDIMENTO prestado?</b>: <br>
        R: {attendance_rating}</p>
        ------------------------------------
        <p><b>De 0 à 10, qual nota você atribui ao seu grau de SATISFAÇÃO com o PRODUTO adquirido?</b>: <br>
        R: {product_satisfaction}</p>
        ------------------------------------
        <p><b>Você RECOMENDARIA a @@@@@ para seus parceiros de negócios ou conhecidos?</b>: <br>
        R: {recommendation}</p>
        ------------------------------------
        <p><b>Tem mais algo que gostaria de compartilhar conosco?</b><br>
        <b>Reclamação, sugestão, agradecimento, etc</b>: <br>
        R: {feedback}</p>
        ___________________________________________________________________________________________________________________
        </body>
        </html>
        """

        # Configurar o servidor de email
        server = smtplib.SMTP(smtp_s, smtp_p)
        server.starttls()  # Ativar o modo TLS
        server.login(sender_email, sender_password)  # Use seu email e senha para autenticação
        
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = recipient_email
        msg['Subject'] = "Feedback de Atendimento e Produto"
        msg.attach(MIMEText(email_body, 'html'))
        
        for attachment_path in attachment_paths:
            if attachment_path:
                attachment = open(attachment_path, 'rb')
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', "attachment; filename= " + os.path.basename(attachment_path))
                msg.attach(part)
        
        server.sendmail(sender_email, recipient_email, msg.as_string())
        server.quit()
        
        # Salvar dados no Excel
        save_to_excel(npedido, current_datetime, cliente_user, Ordem_Compra, Nome_Contato, attendance_rating, product_satisfaction, recommendation, feedback, observaçao)
        
        show_success_message()
    except Exception as e:
        show_error_message(str(e))



def save_to_excel(npedido, datetime, cliente, OC, Nome, attendance, product_satisfaction, recommendation, feedback, observaçao):
    file_path = 'Documento.xlsx'
    sheet_name = 'Cliente'  # Nome da aba onde os dados serão salvos

    # Verifica se o arquivo já existe 
    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
    else:
        workbook = Workbook()
        workbook.remove(workbook.active)  # Remove a aba padrão criada automaticamente
        workbook.create_sheet(title=sheet_name)  # Cria uma nova aba com o nome especificado

    # Seleciona a aba específica
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(title=sheet_name)
    
    # Encontrar a próxima linha vazia verificando as colunas D a P
    next_row = None
    for row in range(1, sheet.max_row + 2):  # Vai até a próxima linha possível
        is_empty = True
        for col in range(4, 17):  # Colunas D a P são 4 a 16
            cell = sheet.cell(row=row, column=col)
            if isinstance(cell, MergedCell) or cell.value not in (None, ""):
                is_empty = False
                break
        if is_empty:
            next_row = row
            break

    if next_row is None:
        next_row = sheet.max_row + 1  # Adiciona na última linha se não encontrou linha vazia

    # Adicionar os dados nas colunas específicas
    sheet[f'D{next_row}'] = npedido
    sheet[f'E{next_row}'] = datetime.split()[0]
    sheet[f'F{next_row}'] = cliente
    sheet[f'G{next_row}'] = OC
    sheet[f'H{next_row}'] = Nome
    sheet[f'I{next_row}'] = 'Sim'
    sheet[f'J{next_row}'] = 'Sim'
    sheet[f'K{next_row}'] = attendance
    sheet[f'L{next_row}'] = product_satisfaction
    sheet[f'M{next_row}'] = recommendation
    sheet[f'N{next_row}'] = feedback
    sheet[f'O{next_row}'] = 'Sim'
    sheet[f'P{next_row}'] = observaçao

    # Salva o arquivo Excel
    workbook.save(file_path)

def show_success_message():
    success_window = tk.Toplevel(window)
    success_window.title("Email Enviado")
    success_window.geometry("300x100")
    success_label = tk.Label(success_window, text="Email enviado com sucesso!", font=("Arial", 12))
    success_label.pack(pady=10)

def show_error_message(error):
    error_window = tk.Toplevel(window)
    error_window.title("Erro ao Enviar Email")
    error_label = tk.Label(error_window, text=f"Erro ao enviar email:\n{error}")
    error_label.pack()
    ok_button = tk.Button(error_window, text="OK", command=lambda: close_windows(error_window))
    ok_button.pack()

def close_windows(window):
    window.destroy()
    window.quit()

def clear_fields():
    email_entry.delete(0, tk.END)
    attendance_combobox.delete(0, tk.END)
    product_combobox.delete(0, tk.END)
    recommendation_combobox.delete(0, tk.END)
    feedback_combobox.delete(0, tk.END)

window = tk.Tk()
window.title("Pesquisa de Satisfação")
window.geometry("700x480")


npedido_label = tk.Label(window, text="Número do Pedido:", width=50, anchor="w", font=("Arial", 12))
npedido_label.grid(row=0, column=0, sticky="w", padx=10, pady=5)
npedido_entry = tk.Entry(window, font=("Arial", 12), width=25)
npedido_entry.place(x=150, y=8)


telefone_label = tk.Label(window, text="Telefone:", width=50, anchor="w", font=("Arial", 12))
telefone_label.grid(row=1, column=0, sticky="w", padx=10, pady=5)
telefone_entry = tk.Entry(window, font=("Arial", 12), width=25)
telefone_entry.place(x=80, y=40)

email_label = tk.Label(window, text="Email:", width=50, anchor="w", font=("Arial", 12))
email_label.grid(row=2, column=0, sticky="w", padx=10, pady=5)
email_entry = tk.Entry(window, font=("Arial", 12), width=40)
email_entry.place(x=60, y=73)


cliente_label = tk.Label(window, text="Empresa:", width=50, anchor="w", font=("Arial", 12))
cliente_label.grid(row=3, column=0, sticky="w", padx=10, pady=5)
cliente_entry = tk.Entry(window, font=("Arial", 12), width=40)
cliente_entry.place(x=85, y=110)

entrevistado_label = tk.Label(window, text="Entrevistado:", width=50, anchor="w", font=("Arial", 12))
entrevistado_label.grid(row=4, column=0, sticky="w", padx=10, pady=5)
entrevistado_entry = tk.Entry(window, font=("Arial", 12), width=40)
entrevistado_entry.place(x=110, y=142)

OC_label = tk.Label(window, text="Ordem de Compra:", width=50, anchor="w", font=("Arial", 12))
OC_label.grid(row=5, column=0, sticky="w", padx=10, pady=5)
OC_entry = tk.Entry(window, font=("Arial", 12), width=40)
OC_entry.place(x=150, y=174)

# Add questions to the GUI
attendance_label = tk.Label(window, text="De 0 à 10, qual nota você atribui ao ATENDIMENTO prestado?:", width=50, anchor="w", font=("Arial", 12))
attendance_label.grid(row=6, column=0, sticky="w", padx=10, pady=5)
attendance_values = [str(i) for i in range(1,11)]  # Lista de valores de 0 a 10
attendance_combobox = ttk.Combobox(window, values=attendance_values, font=("Arial", 12), width=4)
attendance_combobox.place(x=460, y=210)

product_label = tk.Label(window, text="De 0 à 10, qual nota você atribui ao seu grau de \nSATISFAÇÃO com o PRODUTO adquirido?:", width=50, anchor="w", font=("Arial", 11))
product_label.grid(row=7, column=0, sticky="w", padx=10, pady=5)
product_values = [str(i) for i in range(1,11)]   # Exemplos de produtos, substitua pelos seus valores
product_combobox = ttk.Combobox(window, values=product_values, font=("Arial", 12), width=4)
product_combobox.place(x=320, y=265)

recommendation_label = tk.Label(window, text="Você RECOMENDARIA a @ para seus \nparceiros de negócios ou conhecidos?", width=50, anchor="w", font=("Arial", 10))
recommendation_label.grid(row=8, column=0, sticky="w", padx=10, pady=5)
recommendation_values = ['Sim','Não','Outro']
recommendation_combobox = ttk.Combobox(window, values=recommendation_values, font=("Arial", 12), width=10)
recommendation_combobox.place(x=280, y=305)

feedback_label = tk.Label(window, text="Tem mais algo que gostaria de compartilhar conosco?", width=50, anchor="w", font=("Arial", 10))
feedback_label.grid(row=9, column=0, sticky="w", padx=10, pady=5)
feedback_values = ['Sugestão','Agradecimento','Reclamação']
feedback_combobox = ttk.Combobox(window, values=feedback_values, font=("Arial", 11), width=13)
feedback_combobox.place(x=330, y=340)

obs_label = tk.Label(window, text="Observações:", width=50, anchor="w", font=("Arial", 12))
obs_label.grid(row=10, column=0, sticky="w", padx=10, pady=5)
obs_entry = tk.Entry(window, font=("Arial", 10), width=60)
obs_entry.place(x=120, y=375)

send_button = tk.Button(window, text="Enviar Email", command=send_email, font=("Arial", 12, "bold"), bg="lightgreen", width=10)
send_button.grid(row=12, column=0, pady=20)

clear_button = tk.Button(window, text="Apagar", command=clear_fields, font=("Arial", 12, "bold"), bg="lightcoral", width=10)
clear_button.grid(row=12, column=1, pady=20)

window.mainloop()
